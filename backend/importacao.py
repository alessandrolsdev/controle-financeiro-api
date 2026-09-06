# Arquivo: backend/importacao.py
"""Importação de Transações a partir de Planilhas (CSV/XLSX).

Atende à issue #2: permitir que o usuário traga o histórico financeiro de uma
planilha ou extrato bancário sem digitar tudo de novo.

Decisões de engenharia:

- **Sem pandas.** A issue sugeria `pandas` + `openpyxl`, mas o pandas traz
  ~50 MB de dependências (NumPy incluso) para uma tarefa que é essencialmente
  ler linhas e converter tipos. O `csv` da biblioteca padrão e o `openpyxl`
  (que seria necessário de qualquer forma) resolvem com menos superfície de
  ataque e um deploy consideravelmente mais leve.
- **Leitura em modo somente-leitura e streaming.** `openpyxl` com
  `read_only=True` não carrega a planilha inteira na memória — um arquivo
  hostil de 100 MB não derruba o processo.
- **Limites explícitos** de tamanho de arquivo e de linhas, aplicados antes do
  processamento.
- **Neutralização de fórmulas.** Um valor de célula começando com `=`, `+`, `-`
  ou `@` é tratado como texto. Sem isso, exportar os dados de volta para CSV
  reintroduz a fórmula, que o Excel executa ao abrir o arquivo (CSV injection).
- **Erro por linha, não por arquivo.** Uma linha inválida no meio de 500 não
  invalida a importação inteira; o relatório aponta a linha e o motivo.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation

from . import models

# Limites de segurança aplicados antes de qualquer processamento.
TAMANHO_MAXIMO_ARQUIVO = 5 * 1024 * 1024  # 5 MB
MAXIMO_DE_LINHAS = 5000

EXTENSOES_ACEITAS = (".csv", ".xlsx", ".xlsm")

# Nomes aceitos para cada coluna, já normalizados (sem acento, minúsculos).
# A planilha do usuário raramente usa exatamente os nomes que esperamos.
SINONIMOS_DE_COLUNA: dict[str, tuple[str, ...]] = {
    "data": ("data", "date", "dt", "data da compra", "data lancamento",
             "data do lancamento", "data movimento"),
    "descricao": ("descricao", "description", "historico", "historico completo",
                  "lancamento", "titulo", "estabelecimento", "detalhe"),
    "valor": ("valor", "value", "amount", "quantia", "montante", "preco",
              "valor r$", "valor (r$)"),
    "categoria": ("categoria", "category", "classificacao", "tipo de gasto"),
    "tipo": ("tipo", "type", "natureza", "entrada/saida", "d/c"),
    "observacoes": ("observacoes", "observacao", "obs", "notas", "nota",
                    "comentario", "comentarios"),
}

# Heurística de categorização por palavra-chave, usada quando a planilha não
# traz categoria ou traz uma que o usuário não possui.
PALAVRAS_POR_CATEGORIA: dict[str, tuple[str, ...]] = {
    "Transporte": ("uber", "99pop", "99 ", "taxi", "táxi", "cabify", "blablacar",
                   "metro", "metrô", "onibus", "ônibus", "bilhete unico",
                   "combustivel", "combustível", "posto", "shell", "ipiranga",
                   "petrobras", "br mania", "estacionamento", "pedagio",
                   "pedágio", "sem parar", "conectcar"),
    "Alimentação": ("mercado", "supermercado", "atacadao", "atacadão", "assai",
                    "assaí", "carrefour", "pao de acucar", "pão de açúcar",
                    "extra", "big", "sendas", "hortifruti", "padaria",
                    "restaurante", "lanchonete", "ifood", "rappi", "burger",
                    "mcdonald", "subway", "pizzaria", "cafe", "café",
                    "starbucks", "acougue", "açougue"),
    "Moradia": ("aluguel", "condominio", "condomínio", "iptu", "luz", "energia",
                "enel", "cemig", "copel", "light", "agua", "água", "sabesp",
                "gas", "gás", "comgas", "internet", "vivo fibra", "net",
                "claro", "tim", "oi", "telefone"),
    "Saúde": ("farmacia", "farmácia", "drogaria", "droga raia", "drogasil",
              "pacheco", "hospital", "clinica", "clínica", "laboratorio",
              "laboratório", "unimed", "amil", "bradesco saude", "sulamerica",
              "dentista", "consulta", "exame", "psicolog"),
    "Lazer": ("netflix", "spotify", "disney", "hbo", "max ", "prime video",
              "youtube premium", "cinema", "cinemark", "teatro", "show",
              "ingresso", "steam", "playstation", "xbox", "nintendo",
              "academia", "smartfit", "smart fit"),
    "Salário": ("salario", "salário", "pagamento", "provento", "remuneracao",
                "remuneração", "holerite", "folha"),
    "Vendas": ("venda", "vendas", "recebimento", "pix recebido", "transferencia recebida",
               "ted recebida", "deposito", "depósito"),
}

# Prefixos que o Excel interpreta como fórmula ao abrir um CSV.
PREFIXOS_DE_FORMULA = ("=", "+", "-", "@", "\t", "\r")


def normalizar(texto: str) -> str:
    """Normaliza texto para comparação: sem acento, minúsculo, sem espaços extras.

    Args:
        texto (str): O texto de entrada.

    Returns:
        str: O texto normalizado.
    """
    sem_acento = "".join(
        caractere
        for caractere in unicodedata.normalize("NFD", texto)
        if unicodedata.category(caractere) != "Mn"
    )
    return re.sub(r"\s+", " ", sem_acento).strip().lower()


def neutralizar_formula(valor: str) -> str:
    """Impede que um texto da planilha seja executado como fórmula.

    Uma célula com `=cmd|'/c calc'!A1` é apenas texto enquanto está no nosso
    banco, mas vira uma fórmula executável assim que os dados forem exportados
    para CSV e abertos no Excel. O prefixo é neutralizado na entrada.

    Args:
        valor (str): O texto vindo da planilha.

    Returns:
        str: O texto seguro para armazenamento.
    """
    if valor and valor[0] in PREFIXOS_DE_FORMULA:
        return "'" + valor
    return valor


def converter_valor(bruto: object) -> Decimal:
    """Converte um valor monetário em qualquer formato usual para Decimal.

    Trata os dois formatos que aparecem na prática: o brasileiro
    (`1.234,56`) e o anglófono (`1,234.56`), além de símbolo de moeda,
    espaços e parênteses indicando negativo (convenção contábil).

    Args:
        bruto (object): O valor lido da célula.

    Raises:
        ValueError: Se o valor não puder ser interpretado como número.

    Returns:
        Decimal: O valor absoluto, com duas casas decimais.
    """
    if isinstance(bruto, int | float | Decimal):
        valor = Decimal(str(bruto))
    else:
        texto = str(bruto or "").strip()
        if not texto:
            raise ValueError("Valor vazio.")

        # Parênteses indicam negativo na convenção contábil: (150,00).
        negativo = texto.startswith("(") and texto.endswith(")")
        texto = texto.strip("()")

        # Remove símbolo de moeda, espaços (inclusive o não separável) e sinal.
        texto = re.sub(r"[R$\s ]", "", texto, flags=re.IGNORECASE)
        texto = texto.lstrip("+")

        if texto.startswith("-"):
            negativo = True
            texto = texto[1:]

        # Decide qual separador é decimal pelo que aparece por último.
        ultima_virgula = texto.rfind(",")
        ultimo_ponto = texto.rfind(".")

        if ultima_virgula > ultimo_ponto:
            # Formato brasileiro: o ponto é separador de milhar.
            texto = texto.replace(".", "").replace(",", ".")
        else:
            # Formato anglófono: a vírgula é separador de milhar.
            texto = texto.replace(",", "")

        if not texto:
            raise ValueError("Valor vazio.")

        try:
            valor = Decimal(texto)
        except InvalidOperation as erro:
            raise ValueError(f"Valor não numérico: {bruto!r}") from erro

        if negativo:
            valor = -valor

    if not valor.is_finite():
        raise ValueError("Valor não finito.")

    # O sinal do lançamento vem do tipo da categoria, não do valor: um extrato
    # bancário traz despesas como negativas, e é isso que queremos absorver.
    return abs(valor).quantize(Decimal("0.01"))


# Formatos de data testados em ordem. O padrão brasileiro vem primeiro porque
# é o esperado no público-alvo: `03/04/2026` é 3 de abril, não 4 de março.
FORMATOS_DE_DATA = (
    "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d.%m.%Y",
    "%Y-%m-%d", "%Y/%m/%d",
    "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
)


def converter_data(bruto: object) -> datetime:
    """Converte uma data em formato variado para datetime UTC.

    Args:
        bruto (object): O valor lido da célula.

    Raises:
        ValueError: Se a data não puder ser interpretada.

    Returns:
        datetime: A data normalizada em UTC.
    """
    if isinstance(bruto, datetime):
        return (
            bruto.replace(tzinfo=UTC)
            if bruto.tzinfo is None
            else bruto.astimezone(UTC)
        )

    texto = str(bruto or "").strip()
    if not texto:
        raise ValueError("Data vazia.")

    # Remove sufixo de fuso em ISO, que o strptime dos formatos abaixo não lê.
    texto = texto.replace("Z", "").split("+")[0].strip()

    for formato in FORMATOS_DE_DATA:
        try:
            return datetime.strptime(texto, formato).replace(tzinfo=UTC)
        except ValueError:
            continue

    raise ValueError(f"Data em formato não reconhecido: {bruto!r}")


def adivinhar_categoria(descricao: str, tipo_desejado: str) -> str | None:
    """Sugere uma categoria a partir de palavras-chave da descrição.

    Args:
        descricao (str): A descrição do lançamento.
        tipo_desejado (str): 'Gasto' ou 'Receita' — restringe as candidatas.

    Returns:
        str | None: O nome da categoria sugerida, ou None se nada casar.
    """
    texto = normalizar(descricao)

    candidatas = (
        ("Salário", "Vendas")
        if tipo_desejado == models.TIPO_RECEITA
        else ("Transporte", "Alimentação", "Moradia", "Saúde", "Lazer")
    )

    # As palavras são procuradas em fronteira de palavra, e não como substring
    # solta: sem isso "net" (provedor) casa dentro de "netflix", e "oi"
    # (operadora) casa dentro de praticamente qualquer descrição em português.
    for categoria in candidatas:
        for palavra in PALAVRAS_POR_CATEGORIA.get(categoria, ()):
            alvo = normalizar(palavra)
            if not alvo:
                continue
            if re.search(rf"(?<![a-z0-9]){re.escape(alvo)}(?![a-z0-9])", texto):
                return categoria

    return None


def deduzir_tipo(bruto: object, valor_original: str) -> str:
    """Determina se a linha é um gasto ou uma receita.

    Args:
        bruto (object): O conteúdo da coluna de tipo, se houver.
        valor_original (str): O valor como veio na planilha, cujo sinal indica
            a natureza quando não há coluna de tipo.

    Returns:
        str: 'Gasto' ou 'Receita'.
    """
    texto = normalizar(str(bruto or ""))

    if texto:
        if texto in ("c", "credito", "credit", "entrada", "receita", "receitas",
                     "provento", "ganho", "+"):
            return models.TIPO_RECEITA
        if texto in ("d", "debito", "debit", "saida", "despesa", "despesas",
                     "gasto", "gastos", "-"):
            return models.TIPO_GASTO

    # Sem coluna de tipo, o sinal do valor decide — convenção de extrato
    # bancário, em que despesas são negativas.
    limpo = str(valor_original or "").strip()
    if limpo.startswith("-") or (limpo.startswith("(") and limpo.endswith(")")):
        return models.TIPO_GASTO

    return models.TIPO_RECEITA if limpo.startswith("+") else models.TIPO_GASTO


@dataclass
class LinhaImportada:
    """Uma linha da planilha já interpretada.

    Attributes:
        numero (int): Número da linha na planilha, para o relatório de erros.
        descricao (str): Descrição do lançamento.
        valor (Decimal): Valor absoluto.
        data (datetime): Data do lançamento em UTC.
        tipo (str): 'Gasto' ou 'Receita'.
        categoria_sugerida (str | None): Nome da categoria, quando identificada.
        observacoes (str | None): Notas adicionais.
    """

    numero: int
    descricao: str
    valor: Decimal
    data: datetime
    tipo: str
    categoria_sugerida: str | None
    observacoes: str | None


@dataclass
class ErroDeLinha:
    """Uma linha que não pôde ser interpretada.

    Attributes:
        numero (int): Número da linha na planilha.
        motivo (str): Explicação legível do problema.
    """

    numero: int
    motivo: str


@dataclass
class ResultadoDaLeitura:
    """Desfecho da leitura de uma planilha.

    Attributes:
        linhas (list[LinhaImportada]): As linhas interpretadas com sucesso.
        erros (list[ErroDeLinha]): As linhas rejeitadas, com o motivo.
        colunas_detectadas (dict[str, str]): Mapeamento campo → cabeçalho.
    """

    linhas: list[LinhaImportada] = field(default_factory=list)
    erros: list[ErroDeLinha] = field(default_factory=list)
    colunas_detectadas: dict[str, str] = field(default_factory=dict)


class ArquivoInvalidoError(Exception):
    """Sinaliza um arquivo que não pode ser processado como um todo."""


def _mapear_colunas(cabecalhos: list[str]) -> dict[str, int]:
    """Associa cada campo conhecido à posição da coluna correspondente.

    Args:
        cabecalhos (list[str]): Os cabeçalhos lidos da planilha.

    Raises:
        ArquivoInvalidoError: Se faltar uma coluna obrigatória.

    Returns:
        dict[str, int]: Mapeamento campo → índice da coluna.
    """
    normalizados = [normalizar(str(c or "")) for c in cabecalhos]
    mapa: dict[str, int] = {}

    for campo, sinonimos in SINONIMOS_DE_COLUNA.items():
        for indice, cabecalho in enumerate(normalizados):
            if cabecalho in (normalizar(s) for s in sinonimos):
                mapa[campo] = indice
                break

    faltando = [c for c in ("data", "descricao", "valor") if c not in mapa]
    if faltando:
        raise ArquivoInvalidoError(
            "A planilha precisa ter as colunas: "
            + ", ".join(faltando)
            + f". Cabeçalhos encontrados: {', '.join(str(c) for c in cabecalhos if c)}."
        )

    return mapa


def _linhas_do_csv(conteudo: bytes) -> list[list[object]]:
    """Extrai as linhas de um arquivo CSV.

    O delimitador é detectado automaticamente, já que exportações brasileiras
    costumam usar ponto e vírgula em vez de vírgula.

    Args:
        conteudo (bytes): O conteúdo bruto do arquivo.

    Raises:
        ArquivoInvalidoError: Se o arquivo não puder ser decodificado.

    Returns:
        list[list[object]]: As linhas do arquivo.
    """
    texto: str | None = None
    for codificacao in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = conteudo.decode(codificacao)
            break
        except UnicodeDecodeError:
            continue

    if texto is None:
        raise ArquivoInvalidoError("Não foi possível ler a codificação do arquivo.")

    amostra = texto[:8192]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t|")
        delimitador = dialeto.delimiter
    except csv.Error:
        # Sniffer falha em arquivos com uma única coluna; o ponto e vírgula é
        # o padrão mais comum em exportações em português.
        delimitador = ";" if amostra.count(";") > amostra.count(",") else ","

    return [list(linha) for linha in csv.reader(io.StringIO(texto), delimiter=delimitador)]


def _linhas_do_xlsx(conteudo: bytes) -> list[list[object]]:
    """Extrai as linhas da primeira aba de uma planilha XLSX.

    Args:
        conteudo (bytes): O conteúdo bruto do arquivo.

    Raises:
        ArquivoInvalidoError: Se o arquivo não for uma planilha válida.

    Returns:
        list[list[object]]: As linhas da planilha.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as erro:  # pragma: no cover - dependência declarada
        raise ArquivoInvalidoError(
            "Suporte a XLSX indisponível no servidor."
        ) from erro

    try:
        # `read_only` lê em streaming, sem carregar tudo na memória.
        # `data_only` traz o resultado das fórmulas, não o texto delas.
        planilha = load_workbook(
            io.BytesIO(conteudo), read_only=True, data_only=True
        )
    except Exception as erro:  # noqa: BLE001 - qualquer falha é arquivo inválido
        raise ArquivoInvalidoError(
            "O arquivo não é uma planilha XLSX válida ou está corrompido."
        ) from erro

    try:
        aba = planilha.worksheets[0]
        linhas = []
        for indice, linha in enumerate(aba.iter_rows(values_only=True)):
            if indice > MAXIMO_DE_LINHAS:
                break
            linhas.append(list(linha))
        return linhas
    finally:
        planilha.close()


def ler_planilha(nome_do_arquivo: str, conteudo: bytes) -> ResultadoDaLeitura:
    """Interpreta uma planilha CSV ou XLSX de transações.

    Args:
        nome_do_arquivo (str): Nome original, usado para escolher o leitor.
        conteudo (bytes): O conteúdo bruto do arquivo.

    Raises:
        ArquivoInvalidoError: Se o arquivo for grande demais, de formato não
            suportado ou não tiver as colunas obrigatórias.

    Returns:
        ResultadoDaLeitura: As linhas válidas e o relatório de erros.
    """
    if len(conteudo) > TAMANHO_MAXIMO_ARQUIVO:
        raise ArquivoInvalidoError(
            f"O arquivo excede o limite de "
            f"{TAMANHO_MAXIMO_ARQUIVO // (1024 * 1024)} MB."
        )

    if not conteudo:
        raise ArquivoInvalidoError("O arquivo está vazio.")

    nome = (nome_do_arquivo or "").lower()
    if not nome.endswith(EXTENSOES_ACEITAS):
        raise ArquivoInvalidoError(
            f"Formato não suportado. Envie um arquivo {' ou '.join(EXTENSOES_ACEITAS)}."
        )

    linhas = (
        _linhas_do_csv(conteudo)
        if nome.endswith(".csv")
        else _linhas_do_xlsx(conteudo)
    )

    # Ignora linhas totalmente vazias no início do arquivo.
    while linhas and not any(str(c or "").strip() for c in linhas[0]):
        linhas.pop(0)

    if not linhas:
        raise ArquivoInvalidoError("O arquivo não contém dados.")

    if len(linhas) - 1 > MAXIMO_DE_LINHAS:
        raise ArquivoInvalidoError(
            f"O arquivo excede o limite de {MAXIMO_DE_LINHAS} linhas."
        )

    mapa = _mapear_colunas([str(c or "") for c in linhas[0]])
    resultado = ResultadoDaLeitura(
        colunas_detectadas={
            campo: str(linhas[0][indice]) for campo, indice in mapa.items()
        }
    )

    def _celula(linha: list[object], campo: str) -> object:
        """Lê uma célula da linha pelo nome do campo.

        Args:
            linha (list[object]): A linha atual.
            campo (str): O campo lógico desejado.

        Returns:
            object: O conteúdo da célula, ou None se ausente.
        """
        indice = mapa.get(campo)
        if indice is None or indice >= len(linha):
            return None
        return linha[indice]

    for numero, linha in enumerate(linhas[1:], start=2):
        if not any(str(c or "").strip() for c in linha):
            continue  # linha em branco no meio da planilha

        try:
            descricao = str(_celula(linha, "descricao") or "").strip()
            if not descricao:
                raise ValueError("Descrição vazia.")

            valor_bruto = _celula(linha, "valor")
            valor = converter_valor(valor_bruto)
            if valor <= 0:
                raise ValueError("O valor precisa ser maior que zero.")

            data = converter_data(_celula(linha, "data"))

            tipo = deduzir_tipo(_celula(linha, "tipo"), str(valor_bruto or ""))

            categoria = str(_celula(linha, "categoria") or "").strip() or None
            if categoria is None:
                categoria = adivinhar_categoria(descricao, tipo)

            observacoes = str(_celula(linha, "observacoes") or "").strip() or None

            resultado.linhas.append(
                LinhaImportada(
                    numero=numero,
                    descricao=neutralizar_formula(descricao)[:255],
                    valor=valor,
                    data=data,
                    tipo=tipo,
                    categoria_sugerida=(
                        neutralizar_formula(categoria)[:100] if categoria else None
                    ),
                    observacoes=(
                        neutralizar_formula(observacoes)[:2000] if observacoes else None
                    ),
                )
            )
        except ValueError as erro:
            resultado.erros.append(ErroDeLinha(numero=numero, motivo=str(erro)))

    return resultado
